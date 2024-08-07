# # import xlsxwriter module
# import xlsxwriter
#
# # Workbook() takes one, non-optional, argument
# # which is the filename that we want to create.
# workbook = xlsxwriter.Workbook('chart_area.xlsx')
#
# # The workbook object is then used to add new
# # worksheet via the add_worksheet() method.
# worksheet = workbook.add_worksheet()
#
# # Create a new Format object to formats cells
# # in worksheets using add_format() method .
#
# # here we create bold format object .
# bold = workbook.add_format({'bold': 1})
#
# # create a data list .
# headings = ['Number', 'Batch 1', 'Batch 2']
#
# data = [
#     ['2019/12/31', '2020/12/31', '2021/12/31', '2022/12/31', '2023/12/31'],
#     [90,89,20,80,90],
#     [96,78,89,90,98],
#     ['成本/收入'],
#     ['销售费用/收入']
# ]
#
# # Write a row of data starting from 'A1'
# # with bold format .
# worksheet.write_row('B10', data[0])
# worksheet.write_row('B11', data[1])
# worksheet.write_row('B12', data[2])
# worksheet.write_row('A11', data[3])
# worksheet.write_row('A12', data[4])
#
# # Write a column of data starting from
# # worksheet.write_column('B2', data[0])
# # worksheet.write_column('B2', data[1])
# # worksheet.write_column('C2', data[2])
#
# # Create a chart object that can be added
# # to a worksheet using add_chart() method.
#
# # here we create a area chart object .
# chart1 = workbook.add_chart({'type': 'area'})
#
# # Add a data series to a chart
# # using add_series method.
#
# # Configure the first series.
# # = Sheet1 !$A$1 is equivalent to ['Sheet1', 0, 0].
# chart1.add_series({
#     'name': ['Sheet1', 11, 0],
#     'categories': ['Sheet1', 9, 1, 9, 5],
#     'values': ['Sheet1', 11, 1, 11, 5],
# })
#
# # Configure a second series.
# # Note use of alternative syntax to define ranges.
# # [sheetname, first_row, first_col, last_row, last_col].
# chart1.add_series({
#     'name': ['Sheet1', 10, 0],
#     'categories': ['Sheet1', 9, 1, 9, 5],
#     'values': ['Sheet1', 10, 1, 10, 5],
# })
#
# # Add a chart title
# chart1.set_title({'name': 'Results of data analysis'})
#
# # Add x-axis label
# chart1.set_x_axis({'name': 'Test number'})
#
# # Add y-axis label
# chart1.set_y_axis({'name': 'Data length (mm)'})
#
# # Set an Excel chart style.
# chart1.set_legend({'position': 'bottom'})
# chart1.set_style(11)
#
# # add chart to the worksheet
# # the top-left corner of a chart
# # is anchored to cell E2 .
# worksheet.insert_chart('H1', chart1)
#
# # Finally, close the Excel file
# # via the close() method.
# workbook.close()
#
#


import openpyxl
from openpyxl.chart import AreaChart, Reference, Series

# 创建新的工作簿和工作表
workbook = openpyxl.Workbook()
worksheet = workbook.active

# 设置加粗格式
bold_font = openpyxl.styles.Font(bold=True)

# 创建数据列表
headings = ['Number', 'Batch 1', 'Batch 2']
data = [
    ['2019/12/31', '2020/12/31', '2021/12/31', '2022/12/31', '2023/12/31'],
    [90, 89, 20, 80, 90],
    [96, 78, 89, 90, 98],
    ['成本/收入'],
    ['销售费用/收入']
]

# 写入数据到工作表
for col, header in enumerate(headings, start=2):
    cell = worksheet.cell(row=10, column=col)
    cell.value = header
    cell.font = bold_font

# 写入具体的数据
for row, values in enumerate(data, start=11):
    for col, value in enumerate(values, start=2):
        worksheet.cell(row=row, column=col, value=value)

# 创建区域图对象
chart = AreaChart()

# 设置图表的数据系列和类别
for i in range(12, 14):  # 只添加第一和第二组数据
    values = Reference(worksheet, min_col=2, min_row=i, max_col=6, max_row=i)
    series = Series(values, title=data[i-9][0])
    chart.series.append(series)

categories = Reference(worksheet, min_col=2, max_col=6, min_row=11, max_row=11)
chart.set_categories(categories)

# 配置图表标题和轴标签
chart.title = "Results of data analysis"
chart.x_axis.title = "Test number"
chart.y_axis.title = "Data length (mm)"

# 设置图表样式
chart.style = 13  # 预设样式编号13

# 将图表添加到工作表的特定位置
worksheet.add_chart(chart, 'H1')

# 保存工作簿
workbook.save('chart_area_openpyxl.xlsx')

