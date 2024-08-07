import openpyxl
from openpyxl.styles import Alignment
import os
import re


def shift_cell(cell_address, shift_right):
    """平移单元格位置，返回新的单元格位置"""
    # 使用正则表达式分离列字母和行号
    match = re.match(r"([A-Z]+)(\d+)", cell_address)
    if not match:
        raise ValueError("Invalid cell address")

    column_letters = match.group(1)
    row_numbers = match.group(2)

    # 计算新列的索引
    new_column_index = openpyxl.utils.column_index_from_string(column_letters) + shift_right
    new_column_letters = openpyxl.utils.get_column_letter(new_column_index)

    return new_column_letters + row_numbers

# 获取当前文件夹中所有的.xlsx文件
file_names = [f for f in os.listdir('.') if f.endswith('.xlsx')]

for file_name in file_names:
        # 加载现有的Excel文件
    workbook = openpyxl.load_workbook(file_name)

    # 查找包含"资产负债表"的Sheet
    balance_sheets = [workbook[sheet] for sheet in workbook.sheetnames if "资产负债表" in sheet]

    for sheet in balance_sheets:  # 现在sheet是一个Worksheet对象
        # 检查B1单元格的值是否是2024-03-31
        if sheet['B1'].value == '2024-03-31':
            # 删除B列
            sheet.delete_cols(2)  # 第二列是B列

        # 查找2019-12-31所在的列
        column_found = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == '2019-12-31':
                column_found = col[0].column  # 获取2019-12-31所在的列号

        # 如果找到2019-12-31并且该列后还有数据
        if column_found and column_found < sheet.max_column:
            # 删除该列之后的所有列
            sheet.delete_cols(column_found + 1, sheet.max_column - column_found)

    # 保存修改后的Excel文件
    workbook.save(file_name)


# 新建的Sheet名称列表
sheet_names = ["计算公式", "成本收入比率", "其他指标-年报"]



# 遍历每一个文件
for file_name in file_names:
    # 加载现有的Excel文件
    workbook = openpyxl.load_workbook(file_name)

    balance_sheet_name = [s for s in workbook.sheetnames if "资产负债表" in s]

    # 为每个文件添加新的Sheet
    for sheet_name in sheet_names:
        # 创建新Sheet，如果已存在同名Sheet则不创建
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        workbook.create_sheet(sheet_name)

    if balance_sheet_name and "计算公式" in workbook.sheetnames:
        # 获取计算公式sheet
        balance_sheet = workbook[balance_sheet_name[0]]
        calc_sheet = workbook["计算公式"]

        current_assets_cell_address = None
        current_liabilities_cell_address = None
        current_goods_address = None
        current_liabi_address = None
        current_assets_address = None
        current_owners = None
        current_fixeds = None


        # 遍历每一行及其单元格来寻找关键词
        for row in balance_sheet.iter_rows():
            for cell in row:
                if cell.value:
                    if "流动资产合计" == str(cell.value):
                        current_assets_cell_address = cell.coordinate  # 获取单元格坐标
                    elif "流动负债合计" == str(cell.value):
                        current_liabilities_cell_address = cell.coordinate
                    elif "存货" == str(cell.value):
                        current_goods_address = cell.coordinate
                    elif "*负债合计" == str(cell.value):
                        current_liabi_address = cell.coordinate
                    elif "*资产合计" == str(cell.value):
                        current_assets_address = cell.coordinate
                    elif "*所有者权益（或股东权益）合计" == str(cell.value):
                        current_owners = cell.coordinate
                    elif "固定资产合计" == str(cell.value):
                        current_fixeds = cell.coordinate
            # 如果找到了两个合计，就停止搜索
            if current_assets_cell_address and current_liabilities_cell_address and current_goods_address and current_liabi_address and current_assets_address and current_owners and current_fixeds :
                break
        else:  # 如果没有找到合计，再次遍历寻找更一般的术语
            for row in balance_sheet.iter_rows():
                for cell in row:
                    if cell.value:
                        if not current_assets_cell_address and "流动资产" == str(cell.value):
                            current_assets_cell_address = cell.coordinate
                        elif not current_liabilities_cell_address and "流动负债" == str(cell.value):
                            current_liabilities_cell_address = cell.coordinate
                        elif "存货" == str(cell.value):
                            current_goods_address = cell.coordinate
                        elif "*负债合计" == str(cell.value):
                            current_liabi_address = cell.coordinate
                        elif "*资产合计" == str(cell.value):
                            current_assets_address = cell.coordinate
                        elif "所有者权益（或股东权益）合计" == str(cell.value):
                            current_owners = cell.coordinate
                        #     print(f"{file_name} 中流动资产位于单元格: {current_owners}")
                        elif "固定资产合计" == str(cell.value):
                            current_fixeds = cell.coordinate
                # 如果找到了两个非合计，就停止搜索
                if current_assets_cell_address and current_liabilities_cell_address and current_goods_address and current_liabi_address and current_assets_address and current_owners and current_fixeds :
                    break

        # if current_owners:
        #     print(f"{file_name} 中流动资产位于单元格: {current_owners}")
        # if current_fixeds:
        #     print(f"{file_name} 中流动负债位于单元格: {current_fixeds}")


        # 输出找到的单元格位置
        new_assets_addresses = [shift_cell(current_assets_cell_address, i) for i in range(1, 6)]
        new_liabilities_addresses = [shift_cell(current_liabilities_cell_address, i) for i in range(1, 6)]
        new_goods_addresses = [shift_cell(current_goods_address, i) for i in range(1, 6)]
        new_asset = [shift_cell(current_assets_address, i) for i in range(1, 6)]#资产合计
        new_liabi_addresses = [shift_cell(current_liabi_address, i) for i in range(1, 6)]#负债合计
        if current_owners:
            new_owners = [shift_cell(current_owners, i) for i in range(1, 6)]
        if current_fixeds:
            new_fixeds = [shift_cell(current_fixeds, i) for i in range(1, 6)]
        # if new_owners:
        #     print(f"{file_name} 中流动资产位于单元格: {new_owners}")
        # if new_fixeds:
        #     print(f"{file_name} 中流动负债位于单元格: {new_fixeds}")
        #
        #
        calc_sheet.column_dimensions['A'].width = 20
        calc_sheet.column_dimensions['B'].width = 20
        calc_sheet.column_dimensions['C'].width = 20
        calc_sheet.column_dimensions['D'].width = 20
        calc_sheet.column_dimensions['E'].width = 20
        calc_sheet.column_dimensions['F'].width = 20
        # 在A1单元格输入文本
        calc_sheet['A1'] = "偿债能力分析"
        calc_sheet['B1'] = "2023/12/31"
        calc_sheet['C1'] = "2022/12/31"
        calc_sheet['D1'] = "2021/12/31"
        calc_sheet['E1'] = "2020/12/31"
        calc_sheet['F1'] = "2019/12/31"
        calc_sheet['G2'] = "流动比率=流动资产/流动负债"
        calc_sheet['G3'] = "速动比率=速动资产(流动资产-存货)/流动负债"
        calc_sheet['G4'] = "资产负债率=负债总额/资产总额"
        calc_sheet['G5'] = "长期资产适合率=(所有者权益+长期负债)/(固定资产+长期投资)"

        right_alignment = Alignment(horizontal='right')
        cells_to_align_right = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1']

        for cell in cells_to_align_right:
            calc_sheet[cell].alignment = right_alignment

            # 计算每个单元格的值除以基准值
            results = []
            results_goods = []
            results_asset_liab = []
            results_owners_fixeds = []
            for asset_cell, liability_cell, goods_cell, asset, liab, own, fix in zip(new_assets_addresses, new_liabilities_addresses, new_goods_addresses, new_asset, new_liabi_addresses, new_owners, new_fixeds):
                numerator = balance_sheet[asset_cell].value
                denominator = balance_sheet[liability_cell].value
                goods_value = balance_sheet[goods_cell].value
                asset_value = balance_sheet[asset].value
                liab_value = balance_sheet[liab].value
                if current_owners:
                    owns_value = balance_sheet[own].value
                if current_fixeds:
                    fixs_value = balance_sheet[fix].value
                # print(f"{file_name} 中流动资产位于单元格: {new_owners}")
                # print("owns = ", owns_value)
                # print(fixs_value)
                numerator = float(numerator) if numerator is not None else 0.0
                denominator = float(denominator) if denominator is not None else 1.0  # 防止除以0
                result = numerator / denominator
                asset_value = float(asset_value) if asset_value is not None else 0.0
                liab_value = float(liab_value) if liab_value is not None else 0.0
                result_asset_liab_ = liab_value / asset_value
                owns_value = float(owns_value) if owns_value is not None else 0.0
                fixs_value = float(fixs_value) if fixs_value is not None else 0.0
                result_owners_fixeds_ = owns_value / fixs_value
                # print("numerator = ", numerator)

                if goods_value == "--":
                    goods_result = "--"
                else:
                    goods_value = float(goods_value) if goods_value is not None else 0.0


                    goods_result = (numerator - goods_value) / denominator
                results.append(result)
                results_goods.append(goods_result)
                results_asset_liab.append(result_asset_liab_)
                results_owners_fixeds.append(result_owners_fixeds_)

            # 将结果写入计算公式Sheet的B2至F2
            for index, (result, goods_result, asset_liab, owners_fixeds) in enumerate(zip(results, results_goods, results_asset_liab, results_owners_fixeds)):
                # 写入第二行，B2到F2
                cell = calc_sheet[f'{chr(66 + index)}2']  # B2开始
                cell.value = result
                cell.number_format = '0.00%'  # 设置单元格格式为百分比

                # 写入第三行，B3到F3
                goods_cell = calc_sheet[f'{chr(66 + index)}3']  # B3开始
                goods_cell.value = goods_result
                goods_cell.number_format = '0.00%'  # 设置单元格格式为百分比

                asset_liab_cell = calc_sheet[f'{chr(66 + index)}4']  # B4开始
                asset_liab_cell.value = asset_liab
                asset_liab_cell.number_format = '0.00%'  # 设置单元格格式为百分比


                owner_fixed_cell = calc_sheet[f'{chr(66 + index)}5']  # B4开始
                owner_fixed_cell.value = owners_fixeds
                owner_fixed_cell.number_format = '0.00%'  # 设置单元格格式为百分比

        # 保存修改后的Excel文件
        workbook.save(file_name)

