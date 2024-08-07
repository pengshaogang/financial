import copy
import os
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import AreaChart, Reference, Series
from openpyxl.utils import get_column_letter

import xlsxwriter






file_names = [f for f in os.listdir('.') if f.endswith('.xlsx')]

for file_name in file_names:
    workbook = openpyxl.load_workbook(file_name)
    balance_sheets = [workbook[sheet] for sheet in workbook.sheetnames if "资产负债表" in sheet]
    for sheet in balance_sheets:
        if sheet['B1'].value == '2024-03-31':
            sheet.delete_cols(2)

        column_found = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == '2019-12-31':
                column_found = col[0].column

        if column_found and column_found < sheet.max_column:
            sheet.delete_cols(column_found + 1, sheet.max_column - column_found)

    workbook.save(file_name)

sheet_names = ["计算公式", "成本收入比率", "其他指标-年报"]
years = ['B', 'C', 'D', 'E', 'F']

for file_name in file_names:
    workbook = openpyxl.load_workbook(file_name)
    balance_sheet_name = [s for s in workbook.sheetnames if "资产负债表" in s]
    income_sheet_name = [s for s in workbook.sheetnames if "利润表" in s]
    cash_flow_sheet_name = [s for s in workbook.sheetnames if "现金流量表" in s]

    for sheet_name in sheet_names:
        if sheet_name in workbook.sheetnames:
            workbook.remove(workbook[sheet_name])
        workbook.create_sheet(sheet_name)



    if balance_sheet_name and "计算公式" in workbook.sheetnames:
        balance_sheet = workbook[balance_sheet_name[0]]
        calc_sheet = workbook["计算公式"]

        asset_row = None
        liability_row = None
        stock_row = None
        total_liabilities_row = None
        total_assets_row = None
        total_owners_equity_row = None
        accounts_recei_row = None
        total_current_assets_row = None


        for row in balance_sheet.iter_rows(min_row=2, values_only=True):#from row 2 begin, col 1 only
            if row[0] == '流动资产合计':#24
                asset_row = row
            elif row[0] == '流动负债合计':#59
                liability_row = row
            elif row[0] == '存货':
                stock_row = row
            elif row[0] == '*负债合计':
                total_liabilities_row = row
            elif row[0] == '*资产合计':#6
                total_assets_row = row
            elif row[0] == '所有者权益（或股东权益）合计':
                total_owners_equity_row = row
            elif row[0] == '应收账款':
                accounts_recei_row = row

        #...another variable

        results_1 = {'流动比率': []}
        results_2 = {'速动比率': []}
        results_3 = {'资产负债率': []}
        results_4 = {'长期资产适合率': []}
        results_5 = {'资本保值增值率': []}
        results_6 = {'资本积累率': []}
        results_temp_ave_assets = {'平均总资产': []}
        results_temp_ave_eauity = {'平均所有者权益': []}
        results_temp_ave_balance = {'应收帐款平均余额': []}
        results_temp_ave_inventory = {'平均存货': []}
        results_temp_ave_current_assets = {'流动资产平均余额': []}

        #...another result

        if asset_row and liability_row:
            for i in range(1, 6):
                value1 = asset_row[i] if asset_row[i] is not None else '-'
                value2 = liability_row[i] if liability_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_1['流动比率'].append(result)
                    except ZeroDivisionError:
                        results_1['流动比率'].append('-')
                else:
                    results_1['流动比率'].append('-')
        else:
            results_1['流动比率'] = ['-'] * 5
        percent_style_name = 'percent_style'

        # 检查样式是否已存在
        existing_styles = [style for style in workbook.named_styles]
        if percent_style_name not in existing_styles:
            percent_style = NamedStyle(name=percent_style_name, number_format='0.00%')
            workbook.add_named_style(percent_style)
        else:
            percent_style = next(style for style in workbook.named_styles if style == percent_style_name)


        if asset_row and liability_row and stock_row:
            for i in range(1, 6):
                value1 = asset_row[i] if asset_row[i] is not None else '-'
                value2 = liability_row[i] if liability_row[i] is not None else '-'
                value3 = stock_row[i] if stock_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value3 != '-' and value3 != 0:
                    try:
                        # 计算除法
                        result = (float(value1) - float(value3)) / float(value2)
                        results_2['速动比率'].append(result)
                    except ZeroDivisionError:
                        results_2['速动比率'].append('-')
                else:
                    results_2['速动比率'].append('-')
        else:
            results_2['速动比率'] = ['-'] * 5

        if total_liabilities_row and total_assets_row:
            for i in range(1, 6):
                value1 = total_liabilities_row[i] if total_liabilities_row[i] is not None else '-'
                value2 = total_assets_row[i] if total_assets_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_3['资产负债率'].append(result)
                    except ZeroDivisionError:
                        results_3['资产负债率'].append('-')
                else:
                    results_3['资产负债率'].append('-')
        else:
            results_3['资产负债率'] = ['-'] * 5



        results_4['长期资产适合率'] = ['-'] * 5

        if total_assets_row:
            for i in range(1, 6):
                value1 = total_assets_row[i] if total_assets_row[i] is not None else '-'
                if i + 1 < len(total_assets_row):
                    value2 = total_assets_row[i + 1] if total_assets_row[i + 1] is not None else '-'
                else:
                    value2 = '-'
                if value1 != '-' and value2 != '-' and value1 != 0 and value2 != 0:
                    try:
                        # 计算除法
                        result_ave_assets = (float(value1) + float(value2)) / 2.0
                        results_temp_ave_assets['平均总资产'].append(result_ave_assets)
                    except ZeroDivisionError:
                        results_temp_ave_assets['平均总资产'].append('-')
                else:
                    results_temp_ave_assets['平均总资产'].append('-')
        else:
            results_temp_ave_assets['平均总资产'] = ['-'] * 5


        if total_owners_equity_row:
            for i in range(1, 6):
                value1 = total_owners_equity_row[i] if total_owners_equity_row[i] is not None else '-'
                if i + 1 < len(total_owners_equity_row):
                    value2 = total_owners_equity_row[i+1] if total_owners_equity_row[i+1] is not None else '-'
                else:
                    value2 = '-'
                if value1 != '-' and value2 != '-' and value1 != 0 and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        result2 = (float(value1) - float(value2)) / float(value2)
                        result_ave_equity = (float(value1) + float(value2)) / 2.0
                        results_5['资本保值增值率'].append(result)
                        results_6['资本积累率'].append(result2)
                        results_temp_ave_eauity['平均所有者权益'].append(result_ave_equity)
                    except ZeroDivisionError:
                        results_5['资本保值增值率'].append('-')
                        results_6['资本积累率'].append('-')
                        results_temp_ave_eauity['平均所有者权益'].append('-')
                else:
                    results_5['资本保值增值率'].append('-')
                    results_6['资本积累率'].append('-')
                    results_temp_ave_eauity['平均所有者权益'].append('-')
        else:
            results_5['资本保值增值率'] = ['-'] * 5
            results_6['资本积累率'] = ['-'] * 5
            results_temp_ave_eauity['平均所有者权益'] = ['-'] * 5

        if accounts_recei_row:
            for i in range(1, 6):
                value1 = accounts_recei_row[i] if accounts_recei_row[i] is not None else '-'
                if i + 1 < len(accounts_recei_row):
                    value2 = accounts_recei_row[i + 1] if accounts_recei_row[i + 1] is not None else '-'
                else:
                    value2 = '-'
                if value1 != '-' and value2 != '-' and value1 != 0 and value2 != 0:
                    try:
                        # 计算除法
                        result_ave_balance = (float(value1) + float(value2)) / 2.0
                        results_temp_ave_balance['应收帐款平均余额'].append(result_ave_balance)
                    except ZeroDivisionError:
                        results_temp_ave_balance['应收帐款平均余额'].append('-')
                else:
                    results_temp_ave_balance['应收帐款平均余额'].append('-')
        else:
            results_temp_ave_balance['应收帐款平均余额'] = ['-'] * 5

        if stock_row:
            for i in range(1, 6):
                value1 = stock_row[i] if stock_row[i] is not None else '-'
                if i + 1 < len(stock_row):
                    value2 = stock_row[i + 1] if stock_row[i + 1] is not None else '-'
                else:
                    value2 = '-'
                if (value1 != '-' and value1 != '--') and (value2 != '-' and value2 != '--') and value1 != 0 and value2 != 0:
                    try:
                        # 计算除法
                        result_ave_inventory = (float(value1) + float(value2)) / 2.0
                        results_temp_ave_inventory['平均存货'].append(result_ave_inventory)
                    except ZeroDivisionError:
                        results_temp_ave_inventory['平均存货'].append('-')
                else:
                    results_temp_ave_inventory['平均存货'].append('-')
        else:
            results_temp_ave_inventory['平均存货'] = ['-'] * 5

        if asset_row:
            for i in range(1, 6):
                value1 = asset_row[i] if asset_row[i] is not None else '-'
                if i + 1 < len(asset_row):
                    value2 = asset_row[i + 1] if asset_row[i + 1] is not None else '-'
                else:
                    value2 = '-'
                if (value1 != '-' and value1 != '--') and (
                        value2 != '-' and value2 != '--') and value1 != 0 and value2 != 0:
                    try:
                        # 计算除法
                        result_ave_current_assets = (float(value1) + float(value2)) / 2.0
                        results_temp_ave_current_assets['流动资产平均余额'].append(result_ave_current_assets)
                    except ZeroDivisionError:
                        results_temp_ave_current_assets['流动资产平均余额'].append('-')
                else:
                    results_temp_ave_current_assets['流动资产平均余额'].append('-')
        else:
            results_temp_ave_current_assets['流动资产平均余额'] = ['-'] * 5






    if income_sheet_name and "计算公式" in workbook.sheetnames:
        income_sheet = workbook[income_sheet_name[0]]
        calc_sheet = workbook["计算公式"]

        operating_income_row = None
        operating_cost_row = None
        net_profit_row = None
        business_tax_surcharges_row = None
        sell_fee_row = None
        management_fee_row = None
        RD_fee_row = None
        financial_fee_row = None
        total_profit_row = None
        total_operating_income = None
        total_operating_costs = None
        income_tax_row = None



        b1_value = income_sheet['B1'].value
        for row in income_sheet.iter_rows(min_row=2, values_only=True):#from row 2 begin, col 1 only
            if b1_value == '2024-03-31':
                row = row[:1] + row[2:]
            if row[0] == '其中：营业收入':#89
                operating_income_row = row
            elif row[0] == '其中：营业成本':#91
                operating_cost_row = row
            elif row[0] == '五、净利润':#113
                net_profit_row = row
            elif row[0] == '营业税金及附加':#92
                business_tax_surcharges_row = row
            elif row[0] == '销售费用':#93
                sell_fee_row = row
            elif row[0] == '管理费用':#94
                management_fee_row = row
            elif row[0] == '研发费用':#95
                RD_fee_row = row
            elif row[0] == '财务费用':#96
                financial_fee_row = row
            elif row[0] == '四、利润总额':#111
                total_profit_row = row
            elif row[0] == '*营业总收入':#83
                total_operating_income = row
            elif row[0] == '二、营业总成本':#90
                total_operating_costs = row
            elif row[0] == '减：所得税费用':#112
                income_tax_row = row
        #...another variable

        results_income_1 = {'主营业务毛利率': []}
        results_income_2 = {'主营业务净利率': []}
        results_income_3 = {'主营业务成本率': []}
        results_income_4 = {'主营业务税金率': []}
        results_income_5 = {'资产净利率': []}
        results_income_6 = {'净资产收益率': []}
        results_income_7 = {'销售费用率': []}
        results_income_8 = {'管理费用率': []}
        results_income_9 = {'研发费用率': []}
        results_income_10 = {'财务费用率': []}
        results_income_11 = {'成本、费用利润率': []}
        results_temp_operating_income = {'营业总收入': []}
        results_income_12 = {'所得税率': []}
        # results_temp_operating_costs = {'营业总成本': []}

        # ...another result

        if operating_income_row and operating_cost_row:
            for i in range(1, 6):
                value1 = operating_income_row[i] if operating_income_row[i] is not None else '-'
                value2 = operating_cost_row[i] if operating_cost_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = (float(value1) - float(value2)) / float(value1)
                        result_1 = float(value2) / float(value1)
                        results_income_1['主营业务毛利率'].append(result)
                        results_income_3['主营业务成本率'].append(result_1)
                    except ZeroDivisionError:
                        results_income_1['主营业务毛利率'].append('-')
                        results_income_3['主营业务成本率'].append('-')
                else:
                    results_income_1['主营业务毛利率'].append('-')
                    results_income_3['主营业务成本率'].append('-')
        else:
            results_income_1['主营业务毛利率'] = ['-'] * 5
            results_income_3['主营业务成本率'] = ['-'] * 5

        if net_profit_row and operating_income_row:
            for i in range(1, 6):
                value1 = net_profit_row[i] if net_profit_row[i] is not None else '-'
                value2 = operating_income_row[i] if operating_income_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_income_2['主营业务净利率'].append(result)
                    except ZeroDivisionError:
                        results_income_2['主营业务净利率'].append('-')
                else:
                    results_income_2['主营业务净利率'].append('-')
        else:
            results_income_2['主营业务净利率'] = ['-'] * 5

        results_temp_ave_assets['平均总资产'].insert(0,'-')
        if net_profit_row and results_temp_ave_assets['平均总资产']:
            for i in range(1, 6):
                value1 = net_profit_row[i] if net_profit_row[i] is not None else '-'
                value2 = results_temp_ave_assets['平均总资产'][i] if results_temp_ave_assets['平均总资产'][i] != '-' else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_income_5['资产净利率'].append(result)
                    except ZeroDivisionError:
                        results_income_5['资产净利率'].append('-')
                else:
                    results_income_5['资产净利率'].append('-')
        else:
            results_income_5['资产净利率'] = ['-'] * 5
        results_temp_ave_eauity['平均所有者权益'].insert(0,'-')
        if net_profit_row and results_temp_ave_eauity['平均所有者权益']:
            for i in range(1, 6):
                value1 = net_profit_row[i] if net_profit_row[i] is not None else '-'
                value2 = results_temp_ave_eauity['平均所有者权益'][i] if results_temp_ave_eauity['平均所有者权益'][i] != '-' else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_income_6['净资产收益率'].append(result)
                    except ZeroDivisionError:
                        results_income_6['净资产收益率'].append('-')
                else:
                    results_income_6['净资产收益率'].append('-')
        else:
            results_income_6['净资产收益率'] = ['-'] * 5


        if business_tax_surcharges_row and operating_income_row:
            for i in range(1, 6):
                value1 = business_tax_surcharges_row[i] if business_tax_surcharges_row[i] is not None else '-'
                value2 = operating_income_row[i] if operating_income_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_income_4['主营业务税金率'].append(result)
                    except ZeroDivisionError:
                        results_income_4['主营业务税金率'].append('-')
                else:
                    results_income_4['主营业务税金率'].append('-')
        else:
            results_income_4['主营业务税金率'] = ['-'] * 5


        if operating_income_row and sell_fee_row:
            for i in range(1, 6):
                value1 = operating_income_row[i] if operating_income_row[i] is not None else '-'
                value2 = sell_fee_row[i] if sell_fee_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value2) / float(value1)
                        results_income_7['销售费用率'].append(result)
                    except ZeroDivisionError:
                        results_income_7['销售费用率'].append('-')
                else:
                    results_income_7['销售费用率'].append('-')
        else:
            results_income_7['销售费用率'] = ['-'] * 5

        if operating_income_row and management_fee_row:
            for i in range(1, 6):
                value1 = operating_income_row[i] if operating_income_row[i] is not None else '-'
                value2 = management_fee_row[i] if management_fee_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value2) / float(value1)
                        results_income_8['管理费用率'].append(result)
                    except ZeroDivisionError:
                        results_income_8['管理费用率'].append('-')
                else:
                    results_income_8['管理费用率'].append('-')
        else:
            results_income_8['管理费用率'] = ['-'] * 5

        if operating_income_row and RD_fee_row:
            for i in range(1, 6):
                value1 = operating_income_row[i] if operating_income_row[i] is not None else '-'
                value2 = RD_fee_row[i] if RD_fee_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value2) / float(value1)
                        results_income_9['研发费用率'].append(result)
                    except ZeroDivisionError:
                        results_income_9['研发费用率'].append('-')
                else:
                    results_income_9['研发费用率'].append('-')
        else:
            results_income_9['研发费用率'] = ['-'] * 5

        if operating_income_row and financial_fee_row:
            for i in range(1, 6):
                value1 = operating_income_row[i] if operating_income_row[i] is not None else '-'
                value2 = financial_fee_row[i] if financial_fee_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value2) / float(value1)
                        results_income_10['财务费用率'].append(result)
                    except ZeroDivisionError:
                        results_income_10['财务费用率'].append('-')
                else:
                    results_income_10['财务费用率'].append('-')
        else:
            results_income_10['财务费用率'] = ['-'] * 5

        if total_profit_row and operating_cost_row and sell_fee_row and management_fee_row and RD_fee_row:
            for i in range(1, 6):
                value1 = total_profit_row[i] if total_profit_row[i] is not None else '-'
                value2 = operating_cost_row[i] if operating_cost_row[i] is not None else '-'
                value3 = sell_fee_row[i] if sell_fee_row[i] is not None else '-'
                value4 = management_fee_row[i] if management_fee_row[i] is not None else '-'
                value5 = RD_fee_row[i] if RD_fee_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value3 != '-' and value4 != '-' and value5 != '-':
                    try:
                        # 计算除法
                        result = float(value1) / (float(value2) + float(value3) + float(value4) + float(value5))
                        results_income_11['成本、费用利润率'].append(result)
                    except ZeroDivisionError:
                        results_income_11['成本、费用利润率'].append('-')
                else:
                    results_income_11['成本、费用利润率'].append('-')
        else:
            results_income_11['成本、费用利润率'] = ['-'] * 5

        if total_operating_income:
            for i in range(1, 6):
                value1 = total_operating_income[i] if total_operating_income[i] is not None else '-'
                if value1 != '-':
                    try:
                        # 计算除法
                        result = float(value1)
                        results_temp_operating_income['营业总收入'].append(result)
                    except ZeroDivisionError:
                        results_temp_operating_income['营业总收入'].append('-')
                else:
                    results_temp_operating_income['营业总收入'].append('-')
        else:
            results_temp_operating_income['营业总收入'] = ['-'] * 5

        if income_tax_row and total_profit_row:
            for i in range(1, 6):
                value1 = income_tax_row[i] if income_tax_row[i] is not None else '-'
                value2 = total_profit_row[i] if total_profit_row[i] is not None else '-'
                if (value1 != '-' and value1 != '--') and (value2 != '-' and value2 != '--') and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_income_12['所得税率'].append(result)
                    except ZeroDivisionError:
                        results_income_12['所得税率'].append('-')
                else:
                    results_income_12['所得税率'].append('-')
        else:
            results_income_12['所得税率'] = ['-'] * 5




    if cash_flow_sheet_name and "计算公式" in workbook.sheetnames:
        cashflow_sheet = workbook[cash_flow_sheet_name[0]]
        calc_sheet = workbook["计算公式"]

        cash_goods_services_row = None
        net_surplus_row = None
        cash_balance_row = None
        net_cash_row = None

        b1_value = cashflow_sheet['B1'].value
        for row in cashflow_sheet.iter_rows(min_row=2, values_only=True):#from row 2 begin, col 1 only
            if b1_value == '2024-03-31':
                row = row[:1] + row[2:]
            if row[0] == '销售商品、提供劳务收到的现金':#137
                cash_goods_services_row = row
            elif row[0] == '经营活动产生的现金流量净额':#146
                net_surplus_row = row
            elif row[0] == '现金的期末余额':#198
                cash_balance_row = row
            elif row[0] == '*经营活动产生的现金流量净额':#131
                net_cash_row = row



        #...another variable

        results_cash_1 = {'销售收现比': []}
        results_cash_2 = {'营运指数': []}
        results_cash_3 = {'现金比率': []}
        results_cash_4 = {'现金流动负债比': []}
        results_cash_5 = {'现金债务总额比': []}
        results_cash_6 = {'销售现金比率': []}
        results_cash_7 = {'应收帐款周转率(次数)': []}
        results_cash_8 = {'应收帐款周转天数': []}
        results_cash_9 = {'存货周转率': []}
        results_cash_10 = {'存货周转天数': []}
        results_cash_11 = {'流动资产周转次数(率)': []}
        results_cash_12 = {'流动资产周转天数': []}
        results_cash_13 = {'总资产周转率': []}
        results_cash_14 = {'总资产周转天数': []}
        results_cash_15 = {'权益乘数': []}
        results_cash_16 = {'销售净利率': []}
        results_cash_17 = {'权益净利率': []}
        results_cash_18 = {'全部资产现金回收率':[]}

        # ...another result
        results_temp_operating_income['营业总收入'].insert(0,'-')
        # print(results_temp_operating_income)
        if cash_goods_services_row and results_temp_operating_income:
            for i in range(1, 6):
                value1 = cash_goods_services_row[i] if cash_goods_services_row[i] is not None else '-'
                value2 = results_temp_operating_income['营业总收入'][i] if results_temp_operating_income['营业总收入'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_1['销售收现比'].append(result)
                    except ZeroDivisionError:
                        results_cash_1['销售收现比'].append('-')
                else:
                    results_cash_1['销售收现比'].append('-')
        else:
            results_cash_1['销售收现比'] = ['-'] * 5

        if net_surplus_row and cash_goods_services_row:
            for i in range(1, 6):
                value1 = net_surplus_row[i] if net_surplus_row[i] is not None else '-'
                value2 = cash_goods_services_row[i] if cash_goods_services_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_2['营运指数'].append(result)
                    except ZeroDivisionError:
                        results_cash_2['营运指数'].append('-')
                else:
                    results_cash_2['营运指数'].append('-')
        else:
            results_cash_2['营运指数'] = ['-'] * 5

        if liability_row and cash_balance_row:
            for i in range(1, 6):
                value1 = cash_balance_row[i] if cash_balance_row[i] is not None else '-'
                value2 = liability_row[i] if liability_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_3['现金比率'].append(result)
                    except ZeroDivisionError:
                        results_cash_3['现金比率'].append('-')
                else:
                    results_cash_3['现金比率'].append('-')
        else:
            results_cash_3['现金比率'] = ['-'] * 5

        if liability_row and net_cash_row:
            for i in range(1, 6):
                value1 = net_cash_row[i] if net_cash_row[i] is not None else '-'
                value2 = liability_row[i] if liability_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_4['现金流动负债比'].append(result)
                    except ZeroDivisionError:
                        results_cash_4['现金流动负债比'].append('-')
                else:
                    results_cash_4['现金流动负债比'].append('-')
        else:
            results_cash_4['现金流动负债比'] = ['-'] * 5

        if total_liabilities_row and net_cash_row:
            for i in range(1, 6):
                value1 = net_cash_row[i] if net_cash_row[i] is not None else '-'
                value2 = total_liabilities_row[i] if total_liabilities_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_5['现金债务总额比'].append(result)
                    except ZeroDivisionError:
                        results_cash_5['现金债务总额比'].append('-')
                else:
                    results_cash_5['现金债务总额比'].append('-')
        else:
            results_cash_5['现金债务总额比'] = ['-'] * 5

        if total_operating_income and net_cash_row:
            for i in range(1, 6):
                value1 = net_cash_row[i] if net_cash_row[i] is not None else '-'
                value2 = total_operating_income[i] if total_operating_income[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_6['销售现金比率'].append(result)
                    except ZeroDivisionError:
                        results_cash_6['销售现金比率'].append('-')
                else:
                    results_cash_6['销售现金比率'].append('-')
        else:
            results_cash_6['销售现金比率'] = ['-'] * 5

        results_temp_ave_balance['应收帐款平均余额'].insert(0,'-')
        if total_operating_income and results_temp_ave_balance:
            for i in range(1, 6):
                value1 = total_operating_income[i] if total_operating_income[i] is not None else '-'
                value2 = results_temp_ave_balance['应收帐款平均余额'][i] if results_temp_ave_balance['应收帐款平均余额'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        result_turnover_days = 360. / result
                        results_cash_7['应收帐款周转率(次数)'].append(result)
                        results_cash_8['应收帐款周转天数'].append(result_turnover_days)
                    except ZeroDivisionError:
                        results_cash_7['应收帐款周转率(次数)'].append('-')
                        results_cash_8['应收帐款周转天数'].append('-')
                else:
                    results_cash_7['应收帐款周转率(次数)'].append('-')
                    results_cash_8['应收帐款周转天数'].append('-')
        else:
            results_cash_7['应收帐款周转率(次数)'] = ['-'] * 5
            results_cash_8['应收帐款周转天数'] = ['-'] * 5

        results_temp_ave_inventory['平均存货'].insert(0,'-')
        if total_operating_costs and results_temp_ave_inventory:
            for i in range(1, 6):
                value1 = total_operating_costs[i] if total_operating_costs[i] is not None else '-'
                value2 = results_temp_ave_inventory['平均存货'][i] if \
                results_temp_ave_inventory['平均存货'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        result_goods_turnover_days = 360. /result
                        results_cash_9['存货周转率'].append(result)
                        results_cash_10['存货周转天数'].append(result_goods_turnover_days)
                    except ZeroDivisionError:
                        results_cash_9['存货周转率'].append('-')
                        results_cash_10['存货周转天数'].append('-')
                else:
                    results_cash_9['存货周转率'].append('-')
                    results_cash_10['存货周转天数'].append('-')
        else:
            results_cash_9['存货周转率'] = ['-'] * 5
            results_cash_10['存货周转天数'] = ['-'] * 5

        results_temp_ave_current_assets['流动资产平均余额'].insert(0,'-')
        if total_operating_income and results_temp_ave_current_assets:
            for i in range(1, 6):
                value1 = total_operating_income[i] if total_operating_income[i] is not None else '-'
                value2 = results_temp_ave_current_assets['流动资产平均余额'][i] if \
                results_temp_ave_current_assets['流动资产平均余额'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        result_current_assets_turnover_days = 360. /result
                        results_cash_11['流动资产周转次数(率)'].append(result)
                        results_cash_12['流动资产周转天数'].append(result_current_assets_turnover_days)
                    except ZeroDivisionError:
                        results_cash_11['流动资产周转次数(率)'].append('-')
                        results_cash_12['流动资产周转天数'].append('-')
                else:
                    results_cash_11['流动资产周转次数(率)'].append('-')
                    results_cash_12['流动资产周转天数'].append('-')
        else:
            results_cash_11['流动资产周转次数(率)'] = ['-'] * 5
            results_cash_12['流动资产周转天数'] = ['-'] * 5

        if total_operating_income and results_temp_ave_assets:
            for i in range(1, 6):
                value1 = total_operating_income[i] if total_operating_income[i] is not None else '-'
                value2 = results_temp_ave_assets['平均总资产'][i] if \
                    results_temp_ave_assets['平均总资产'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        result_total_assets_turnover_days = 360. / result
                        results_cash_13['总资产周转率'].append(result)
                        results_cash_14['总资产周转天数'].append(result_total_assets_turnover_days)
                    except ZeroDivisionError:
                        results_cash_13['总资产周转率'].append('-')
                        results_cash_14['总资产周转天数'].append('-')
                else:
                    results_cash_13['总资产周转率'].append('-')
                    results_cash_14['总资产周转天数'].append('-')
        else:
            results_cash_13['总资产周转率'] = ['-'] * 5
            results_cash_14['总资产周转天数'] = ['-'] * 5

        results_3_copy = copy.deepcopy(results_3)
        results_3_copy['资产负债率'].insert(0,'-')
        if results_3_copy:
            for i in range(1, 6):
                value1 = results_3_copy['资产负债率'][i] if results_3_copy['资产负债率'][i] is not None else '-'
                if value1 != '-':
                    try:
                        # 计算除法
                        result = 1.0 / (1.0 - value1)
                        results_cash_15['权益乘数'].append(result)
                    except ZeroDivisionError:
                        results_cash_15['权益乘数'].append('-')
                else:
                    results_cash_15['权益乘数'].append('-')
        else:
            results_cash_15['权益乘数'] = ['-'] * 5

        if net_profit_row and total_operating_income:
            for i in range(1, 6):
                value1 = net_profit_row[i] if net_profit_row[i] is not None else '-'
                value2 = total_operating_income[i] if \
                    total_operating_income[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_16['销售净利率'].append(result)
                    except ZeroDivisionError:
                        results_cash_16['销售净利率'].append('-')
                else:
                    results_cash_16['销售净利率'].append('-')
        else:
            results_cash_16['销售净利率'] = ['-'] * 5


        results_cash_13_copy = copy.deepcopy(results_cash_13)
        results_cash_15_copy = copy.deepcopy(results_cash_15)
        results_cash_16_copy = copy.deepcopy(results_cash_16)
        results_cash_13_copy['总资产周转率'].insert(0,'-')
        results_cash_15_copy['权益乘数'].insert(0,'-')
        results_cash_16_copy['销售净利率'].insert(0,'-')
        if results_cash_13_copy['总资产周转率'] and results_cash_15_copy['权益乘数'] and results_cash_16_copy['销售净利率']:
            for i in range(1, 6):
                value1 = results_cash_13_copy['总资产周转率'][i] if results_cash_13_copy['总资产周转率'][i] is not None else '-'
                value2 = results_cash_15_copy['权益乘数'][i] if results_cash_15_copy['权益乘数'][i] is not None else '-'
                value3 = results_cash_16_copy['销售净利率'][i] if results_cash_16_copy['销售净利率'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value3 != '-':
                    try:
                        # 计算除法
                        result = float(value1) * float(value2) * float(value3)
                        results_cash_17['权益净利率'].append(result)
                    except ZeroDivisionError:
                        results_cash_17['权益净利率'].append('-')
                else:
                    results_cash_17['权益净利率'].append('-')
        else:
            results_cash_17['权益净利率'] = ['-'] * 5

        if net_cash_row and total_assets_row:
            for i in range(1, 6):
                value1 = net_cash_row[i] if net_cash_row[i] is not None else '-'
                value2 = total_assets_row[i] if \
                    total_assets_row[i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != 0:
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cash_18['全部资产现金回收率'].append(result)
                    except ZeroDivisionError:
                        results_cash_18['全部资产现金回收率'].append('-')
                else:
                    results_cash_18['全部资产现金回收率'].append('-')
        else:
            results_cash_18['全部资产现金回收率'] = ['-'] * 5

    if income_sheet_name and "成本收入比率" in workbook.sheetnames:
        income_sheet = workbook[income_sheet_name[0]]
        cost_income_sheet = workbook["成本收入比率"]



        results_cost_income_1 = {'营业总成本': []}
        results_cost_income_2 = {'营业税金及附加': []}
        results_cost_income_3 = {'销售费用': []}
        results_cost_income_4 = {'管理费用': []}
        results_cost_income_5 = {'研发费用': []}
        results_cost_income_6 = {'毛利润': []}
        results_cost_income_7 = {'成本/收入': []}
        results_cost_income_8 = {'销售费用/收入': []}
        results_cost_income_9 = {'管理费用/收入': []}
        results_cost_income_10 = {'研发费用/收入': []}
        results_cost_income_11 = {'毛利润率': []}

        results_temp_operating_income_copy = copy.deepcopy(results_temp_operating_income)
        del results_temp_operating_income_copy['营业总收入'][0]

        if total_operating_costs:
            for i in range(1, 6):
                value1 = total_operating_costs[i] if total_operating_costs[i] is not None else '-'
                if value1 != '-'  and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value1)
                        results_cost_income_1['营业总成本'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_1['营业总成本'].append('-')
                else:
                    results_cost_income_1['营业总成本'].append('-')
        else:
            results_cost_income_1['营业总成本'] = ['-'] * 5

        if business_tax_surcharges_row:
            for i in range(1, 6):
                value1 = business_tax_surcharges_row[i] if business_tax_surcharges_row[i] is not None else '-'
                if value1 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value1)
                        results_cost_income_2['营业税金及附加'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_2['营业税金及附加'].append('-')
                else:
                    results_cost_income_2['营业税金及附加'].append('-')
        else:
            results_cost_income_2['营业税金及附加'] = ['-'] * 5

        if sell_fee_row:
            for i in range(1, 6):
                value1 = sell_fee_row[i] if sell_fee_row[i] is not None else '-'
                if value1 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value1)
                        results_cost_income_3['销售费用'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_3['销售费用'].append('-')
                else:
                    results_cost_income_3['销售费用'].append('-')
        else:
            results_cost_income_3['销售费用'] = ['-'] * 5

        if management_fee_row:
            for i in range(1, 6):
                value1 = management_fee_row[i] if management_fee_row[i] is not None else '-'
                if value1 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value1)
                        results_cost_income_4['管理费用'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_4['管理费用'].append('-')
                else:
                    results_cost_income_4['管理费用'].append('-')
        else:
            results_cost_income_4['管理费用'] = ['-'] * 5

        if RD_fee_row:
            for i in range(1, 6):
                value1 = RD_fee_row[i] if RD_fee_row[i] is not None else '-'
                if value1 != '-' and value1 != 0:
                    try:
                        # 计算除法
                        result = float(value1)
                        results_cost_income_5['研发费用'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_5['研发费用'].append('-')
                else:
                    results_cost_income_5['研发费用'].append('-')
        else:
            results_cost_income_5['研发费用'] = ['-'] * 5


        if results_temp_operating_income_copy and results_cost_income_1 and results_cost_income_2 and results_cost_income_3 and results_cost_income_4 and results_cost_income_5:
            for i in range(5):
                value1 = results_temp_operating_income_copy['营业总收入'][i] if results_temp_operating_income_copy['营业总收入'][i] is not None else '-'
                value2 = results_cost_income_1['营业总成本'][i] if results_cost_income_1['营业总成本'][i] is not None else '-'
                value3 = results_cost_income_2['营业税金及附加'][i] if results_cost_income_2['营业税金及附加'][i] is not None else '-'
                value4 = results_cost_income_3['销售费用'][i] if results_cost_income_3['销售费用'][i] is not None else '-'
                value5 = results_cost_income_4['管理费用'][i] if results_cost_income_4['管理费用'][i] is not None else '-'
                value6 = results_cost_income_5['研发费用'][i] if results_cost_income_5['研发费用'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value3 != '-' and value4 != '-' and value5 != '-' and value6 != '-':
                    try:
                        # 计算除法
                        result = float(value1) - float(value2) - float(value3) - float(value4) - float(value5) - float(value6)
                        results_cost_income_6['毛利润'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_6['毛利润'].append('-')
                else:
                    results_cost_income_6['毛利润'].append('-')
        else:
            results_cost_income_6['毛利润'] = ['-'] * 5


        if results_temp_operating_income_copy and results_cost_income_1:
            for i in range(5):
                value1 = results_cost_income_1['营业总成本'][i] if results_cost_income_1['营业总成本'][i] is not None else '-'
                value2 = results_temp_operating_income_copy['营业总收入'][i] if results_temp_operating_income_copy['营业总收入'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != '0':
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cost_income_7['成本/收入'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_7['成本/收入'].append('-')
                else:
                    results_cost_income_7['成本/收入'].append('-')
        else:
            results_cost_income_7['成本/收入'] = ['-'] * 5

        if results_temp_operating_income_copy and results_cost_income_3:
            for i in range(5):
                value1 = results_cost_income_3['销售费用'][i] if results_cost_income_3['销售费用'][i] is not None else '-'
                value2 = results_temp_operating_income_copy['营业总收入'][i] if results_temp_operating_income_copy['营业总收入'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != '0':
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cost_income_8['销售费用/收入'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_8['销售费用/收入'].append('-')
                else:
                    results_cost_income_8['销售费用/收入'].append('-')
        else:
            results_cost_income_8['销售费用/收入'] = ['-'] * 5


        if results_temp_operating_income_copy and results_cost_income_4:
            for i in range(5):
                value1 = results_cost_income_4['管理费用'][i] if results_cost_income_4['管理费用'][i] is not None else '-'
                value2 = results_temp_operating_income_copy['营业总收入'][i] if results_temp_operating_income_copy['营业总收入'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != '0':
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cost_income_9['管理费用/收入'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_9['管理费用/收入'].append('-')
                else:
                    results_cost_income_9['管理费用/收入'].append('-')
        else:
            results_cost_income_9['管理费用/收入'] = ['-'] * 5

        if results_temp_operating_income_copy and results_cost_income_5:
            for i in range(5):
                value1 = results_cost_income_5['研发费用'][i] if results_cost_income_5['研发费用'][i] is not None else '-'
                value2 = results_temp_operating_income_copy['营业总收入'][i] if results_temp_operating_income_copy['营业总收入'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != '0':
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cost_income_10['研发费用/收入'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_10['研发费用/收入'].append('-')
                else:
                    results_cost_income_10['研发费用/收入'].append('-')
        else:
            results_cost_income_10['研发费用/收入'] = ['-'] * 5


        if results_temp_operating_income_copy and results_cost_income_6:
            for i in range(5):
                value1 = results_cost_income_6['毛利润'][i] if results_cost_income_6['毛利润'][i] is not None else '-'
                value2 = results_temp_operating_income_copy['营业总收入'][i] if results_temp_operating_income_copy['营业总收入'][i] is not None else '-'
                if value1 != '-' and value2 != '-' and value2 != '0':
                    try:
                        # 计算除法
                        result = float(value1) / float(value2)
                        results_cost_income_11['毛利润率'].append(result)
                    except ZeroDivisionError:
                        results_cost_income_11['毛利润率'].append('-')
                else:
                    results_cost_income_11['毛利润率'].append('-')
        else:
            results_cost_income_11['毛利润率'] = ['-'] * 5

        # years_plot = ['2019/12/31', '2020/12/31', '2021/12/31', '2022/12/31', '2023/12/31']
        # results_cost_income_7_copy = {'成本/收入': []}
        # results_cost_income_7_copy['成本/收入'] = results_cost_income_7['成本/收入'][::-1]
        #
        # data_plot = [
        #     [years_plot[0],results_cost_income_7_copy['成本/收入'][0]],
        #     [years_plot[1], results_cost_income_7_copy['成本/收入'][1]],
        #     [years_plot[2], results_cost_income_7_copy['成本/收入'][2]],
        #     [years_plot[3], results_cost_income_7_copy['成本/收入'][3]],
        #     [years_plot[4], results_cost_income_7_copy['成本/收入'][4]],
        # ]



            # chart.grouping = 'stacked'  # 设置图表为堆叠面积图
            # chart.title = "成本收入比率"
            #
            #
            #
            # categories = Reference(cost_income_sheet, min_col=2, min_row=10, max_col=6, max_row=10)
            # chart.set_categories(categories)
            #
            # for row in cost_income_sheet.iter_rows(min_row=categories.min_row, max_row=categories.max_row, min_col=categories.min_col,
            #                         max_col=categories.max_col):
            #     values = [cell.value for cell in row]
            #     print(values)
            #
            # for i in range(11, 12):
            #     chart_data = Reference(cost_income_sheet, min_col=2, min_row=i, max_col=6, max_row=i)
            #     chart.add_data(chart_data, titles_from_data=True)
            #
            #
            #
            # chart.title = "成本收入比率"
            #
            # cost_income_sheet.add_chart(chart, "H1")






























        #...another judgement




    calc_sheet.column_dimensions['A'].width = 20
    calc_sheet.column_dimensions['B'].width = 20
    calc_sheet.column_dimensions['C'].width = 20
    calc_sheet.column_dimensions['D'].width = 20
    calc_sheet.column_dimensions['E'].width = 20
    calc_sheet.column_dimensions['F'].width = 20
    calc_sheet.column_dimensions['G'].width = 20

    cost_income_sheet.column_dimensions['A'].width = 20
    cost_income_sheet.column_dimensions['B'].width = 20
    cost_income_sheet.column_dimensions['C'].width = 20
    cost_income_sheet.column_dimensions['D'].width = 20
    cost_income_sheet.column_dimensions['E'].width = 20
    cost_income_sheet.column_dimensions['F'].width = 20
    cost_income_sheet.column_dimensions['G'].width = 20

    # 在A1单元格输入文本
    calc_sheet['A1'] = "偿债能力分析"
    calc_sheet['B1'] = "2023/12/31"
    calc_sheet['C1'] = "2022/12/31"
    calc_sheet['D1'] = "2021/12/31"
    calc_sheet['E1'] = "2020/12/31"
    calc_sheet['F1'] = "2019/12/31"
    calc_sheet['G1'] = "取数"
    calc_sheet['G2'] = "流动比率=流动资产/流动负债"
    calc_sheet['G3'] = "速动比率=速动资产(流动资产-存货)/流动负债"
    calc_sheet['G4'] = "资产负债率=负债总额/资产总额"
    calc_sheet['G5'] = "长期资产适合率=(所有者权益+长期负债)/(固定资产+长期投资)"

    calc_sheet['A6'] = "资本状况分析"
    calc_sheet['B6'] = "2023/12/31"
    calc_sheet['C6'] = "2022/12/31"
    calc_sheet['D6'] = "2021/12/31"
    calc_sheet['E6'] = "2020/12/31"
    calc_sheet['F6'] = "2019/12/31"
    calc_sheet['G7'] = "资本保值增值率=扣除客观因素后的年末所有者权益/年初所有者权益"
    calc_sheet['G8'] = "资本积累率=本年所有者权益增长额/年初所有者权益"

    calc_sheet['A9'] = "盈利能力分析"
    calc_sheet['B9'] = "2023/12/31"
    calc_sheet['C9'] = "2022/12/31"
    calc_sheet['D9'] = "2021/12/31"
    calc_sheet['E9'] = "2020/12/31"
    calc_sheet['F9'] = "2019/12/31"
    calc_sheet['G10'] = "主营业务毛利率=(主营收入-主营成本)/业务收入"
    calc_sheet['G11'] = "主营业务净利率=净利润/主营业务收入"
    calc_sheet['G12'] = "主营业务成本率=主营业务成本/主营业务收入"
    calc_sheet['G13'] = "主营业务税金率=主营业务税金及附加/主营业务收入"
    calc_sheet['G14'] = "资产净利率=税后净利/平均总资产=主营业务净利率*总资产周转率"
    calc_sheet['G15'] = "净资产收益率=113净利润/平均所有者权益/2"
    calc_sheet['G16'] = "销售费用率=93销售费用/89主营业务收入"
    calc_sheet['G17'] = "管理费用率=94管理费用/89主营业务收入"
    calc_sheet['G18'] = "研发费用率=95研发费用/89主营业务收入"
    calc_sheet['G19'] = "财务费用率=96财务费用/89主营业务收入"
    calc_sheet['G20'] = "成本、费用利润率=111利润总额/(91主营业务成本+93+94+95期间费用)"
    calc_sheet['G21'] = "销售收现比=137销售收现/83销售额"
    calc_sheet['G22'] = "营运指数=146经营现金净流量/137经营所得现金"
    calc_sheet['G23'] = "现金比率=198现金余额/59流动负债"
    calc_sheet['G24'] = "现金流动负债比=131经营活动净现金流量/59流动负债"
    calc_sheet['G25'] = "现金债务总额比=131经营活动净现金流量/7总负债"
    calc_sheet['G26'] = "销售现金比率=131经营现金净流量/83销售额"

    calc_sheet['A27'] = "注：财务报表只显示公司的总销售额，不会单独列出赊销收入。赊销收入净额是一个会计概念，需要从总销售额中扣除一些项目（如销货退回、销货折扣与折让等）来得到。"
    calc_sheet['A28'] = "这细节会在财务报表的附注中提供，不是在财务报表正文中显示。了解赊销收入净额的信息，需要查看财务报表附注或相关财务报告。"
    calc_sheet['A29'] = "所以赊销收入净额如果无法找到附注的话，先暂用营业收入金额来替代。"

    calc_sheet['A30'] = "营运能力分析"
    calc_sheet['B30'] = "2023/12/31"
    calc_sheet['C30'] = "2022/12/31"
    calc_sheet['D30'] = "2021/12/31"
    calc_sheet['E30'] = "2020/12/31"
    calc_sheet['F30'] = "2019/12/31"

    calc_sheet['G31'] = "应收帐款周转率(次数)=83赊销收入净额/应收帐款平均余额（C15+D15）/2"
    calc_sheet['G32'] = "应收帐款周转天数=360天数/应收帐款周转次数=计算期天数*应收帐款平均余额/赊销收入净额"
    calc_sheet['G33'] = "存货周转率=90营业成本/平均存货（C20+D20）/2"
    calc_sheet['G34'] = "存货周转天数=计算期天数(年360天)/存货周转率"
    calc_sheet['G35'] = "流动资产周转次数(率)=83销售收入净额/流动资产平均余额（C24+D24）/2"
    calc_sheet['G36'] = "流动资产周转天数=360计算期天数/流动资产周转次数"
    calc_sheet['G37'] = "总资产周转率=83销售收入/平均资产总额(C6+D6)/2"
    calc_sheet['G38'] = "总资产周转天数=360计算期天数/总资产周转率"

    calc_sheet['A39'] = "利息支出数据取不到，是否可以取到年报数据？"
    calc_sheet['G39'] = "已获利息倍数=(税前利润总额+利息支出)/利息支出"
    calc_sheet['G40'] = "总资产报酬率=(利润总额+利息支出)/平均资产总额"

    calc_sheet['A42'] = "其他财务报表分析所用公式"
    calc_sheet['A43'] = "(一) 杜帮财务分析体系"
    calc_sheet['B43'] = "2023/12/31"
    calc_sheet['C43'] = "2022/12/31"
    calc_sheet['D43'] = "2021/12/31"
    calc_sheet['E43'] = "2020/12/31"
    calc_sheet['F43'] = "2019/12/31"

    calc_sheet['G44'] = "权益乘数=1÷(1-资产负债率)"
    calc_sheet['G45'] = "销售净利率=113净利润/83销售收入"
    calc_sheet['G46'] = "总资产周转率=83销售收入/平均资产总额(C6+D6)/2"
    calc_sheet['G47'] = "权益净利率=资产净利率×权益乘数=销售净利率×资产周转率×权益乘数"

    calc_sheet['A50'] = "(二) 上市公司财务比率"
    calc_sheet['B50'] = "2023/12/31"
    calc_sheet['C50'] = "2022/12/31"
    calc_sheet['D50'] = "2021/12/31"
    calc_sheet['E50'] = "2020/12/31"
    calc_sheet['F50'] = "2019/12/31"

    calc_sheet['G50'] = "以下数据可以直接抓取"
    calc_sheet['G51'] = "每股收益=净利润÷年末普通股份总数=(净利润-优先股股利)÷(年度股份总数-年度末优先股数)"
    calc_sheet['G52'] = "市盈率(倍数)=普通股每股市价÷普通股每股收益"
    calc_sheet['G53'] = "每股股利=股利总额÷年末普通股股份总数"
    calc_sheet['G54'] = "股票获利率=普通股每股股利÷普通股每股市价×100%"
    calc_sheet['G55'] = "股利支付率=(普通股每股股利÷普通股每股净收益)×100%"
    calc_sheet['G56'] = "股利保障倍数=普通股每股净收益÷普通股每股股利=1÷股利支付率"
    calc_sheet['G57'] = "每股净资产=年度末股东权益÷年度末普通股数"
    calc_sheet['G58'] = "市净率(倍数)=每股市价÷每股净资产"

    calc_sheet['A61'] = "(三) 现金流量分析"
    calc_sheet['A62'] = "1、流动性分析"
    calc_sheet['B62'] = "2023/12/31"
    calc_sheet['C62'] = "2022/12/31"
    calc_sheet['D62'] = "2021/12/31"
    calc_sheet['E62'] = "2020/12/31"
    calc_sheet['F62'] = "2019/12/31"
    calc_sheet['G63'] = "现金流动负债比=131经营现金流量净额÷59流动负债"
    calc_sheet['G64'] = "现金债务总额比=131经营现金流量净额÷7期末负债总额"

    calc_sheet['A66'] = "2、获取现金能力分析"
    calc_sheet['B66'] = "2023/12/31"
    calc_sheet['C66'] = "2022/12/31"
    calc_sheet['D66'] = "2021/12/31"
    calc_sheet['E66'] = "2020/12/31"
    calc_sheet['F66'] = "2019/12/31"
    calc_sheet['G67'] = "销售现金比率=131经营现金流量净额÷88销售额"
    calc_sheet['G68'] = "每股经营现金流量净额=经营现金流量净额÷普通股股数"
    calc_sheet['G69'] = "全部资产现金回收率=131经营现金流量净额÷6全部资产"

    calc_sheet['A71'] = "3、财务弹性分析"
    calc_sheet['B71'] = "2023/12/31"
    calc_sheet['C71'] = "2022/12/31"
    calc_sheet['D71'] = "2021/12/31"
    calc_sheet['E71'] = "2020/12/31"
    calc_sheet['F71'] = "2019/12/31"
    calc_sheet['G72'] = "现金满足投资比率=近5年经营现金流量净额之和÷近5年资本支出、存货增加、现金股利之和"
    calc_sheet['G73'] = "现金股利保障倍数=每股经营现金流量净额÷每股现金股"

    calc_sheet['A75'] = "其他:"
    calc_sheet['G76'] = "主营业务税金率=92主营业务税金及附加/89主营业务收入"
    calc_sheet['G77'] = "所得税率=112所得税/111利润总额"

    cost_income_sheet['A1'] = "科目\年度"
    cost_income_sheet['B1'] = "2023/12/31"
    cost_income_sheet['C1'] = "2022/12/31"
    cost_income_sheet['D1'] = "2021/12/31"
    cost_income_sheet['E1'] = "2020/12/31"
    cost_income_sheet['F1'] = "2019/12/31"
    cost_income_sheet['G1'] = "五年均值"



    cost_income_sheet['A2'] = "一、营业总收入(万元)"
    cost_income_sheet['A3'] = "二、营业总成本(万元)"
    cost_income_sheet['A4'] = "营业税金及附加(万元)"
    cost_income_sheet['A5'] = "销售费用(万元)"
    cost_income_sheet['A6'] = "管理费用(万元)"
    cost_income_sheet['A7'] = "研发费用(万元)"
    cost_income_sheet['A8'] = "毛利润=营业收入-营业成本-管销研费用-营业税金及附加"

    cost_income_sheet['G2'] = float(sum(results_temp_operating_income_copy['营业总收入'])) / float(len(results_temp_operating_income_copy['营业总收入'])) if all(isinstance(item, (float, int)) for item in results_temp_operating_income_copy['营业总收入']) else '-'
    cost_income_sheet['G3'] = float(sum(results_cost_income_1['营业总成本'])) / float(len(results_cost_income_1['营业总成本'])) if all(isinstance(item, (float, int)) for item in results_cost_income_1['营业总成本']) else '-'
    cost_income_sheet['G4'] = float(sum(results_cost_income_2['营业税金及附加'])) / float(len(results_cost_income_2['营业税金及附加'])) if all(isinstance(item, (float, int)) for item in results_cost_income_2['营业税金及附加']) else '-'
    cost_income_sheet['G5'] = float(sum(results_cost_income_3['销售费用'])) / float(len(results_cost_income_3['销售费用'])) if all(isinstance(item, (float, int)) for item in results_cost_income_3['销售费用']) else '-'
    cost_income_sheet['G6'] = float(sum(results_cost_income_4['管理费用'])) / float(len(results_cost_income_4['管理费用'])) if all(isinstance(item, (float, int)) for item in results_cost_income_4['管理费用']) else '-'
    cost_income_sheet['G7'] = float(sum(results_cost_income_5['研发费用'])) / float(len(results_cost_income_5['研发费用'])) if all(isinstance(item, (float, int)) for item in results_cost_income_5['研发费用']) else '-'
    cost_income_sheet['G8'] = float(sum(results_cost_income_6['毛利润'])) / float(len(results_cost_income_6['毛利润'])) if all(isinstance(item, (float, int)) for item in results_cost_income_6['毛利润']) else '-'

    cost_income_sheet['G11'] = float(sum(results_cost_income_7['成本/收入'])) / float(len(results_cost_income_7['成本/收入'])) if all(isinstance(item, (float, int)) for item in results_cost_income_7['成本/收入']) else '-'
    cost_income_sheet['G12'] = float(sum(results_cost_income_8['销售费用/收入'])) / float(len(results_cost_income_8['销售费用/收入'])) if all(isinstance(item, (float, int)) for item in results_cost_income_8['销售费用/收入']) else '-'
    cost_income_sheet['G13'] = float(sum(results_cost_income_9['管理费用/收入'])) / float(len(results_cost_income_9['管理费用/收入'])) if all(isinstance(item, (float, int)) for item in results_cost_income_9['管理费用/收入']) else '-'
    cost_income_sheet['G14'] = float(sum(results_cost_income_10['研发费用/收入'])) / float(len(results_cost_income_10['研发费用/收入'])) if all(isinstance(item, (float, int)) for item in results_cost_income_10['研发费用/收入']) else '-'
    cost_income_sheet['G15'] = float(sum(results_cost_income_11['毛利润率'])) / float(len(results_cost_income_11['毛利润率'])) if all(isinstance(item, (float, int)) for item in results_cost_income_11['毛利润率']) else '-'



    cost_income_sheet['A10'] = "比率"
    cost_income_sheet['B10'].number_format = 'YYYY/MM/DD'
    cost_income_sheet['B10'] = datetime.strptime("2023/12/31", '%Y/%m/%d')

    cost_income_sheet['C10'] = datetime.strptime("2022/12/31", '%Y/%m/%d')
    cost_income_sheet['C10'].number_format = 'YYYY/MM/DD'
    cost_income_sheet['D10'] = datetime.strptime("2021/12/31", '%Y/%m/%d')
    cost_income_sheet['D10'].number_format = 'YYYY/MM/DD'
    cost_income_sheet['E10'] = datetime.strptime("2020/12/31", '%Y/%m/%d')
    cost_income_sheet['E10'].number_format = 'YYYY/MM/DD'
    cost_income_sheet['F10'] = datetime.strptime("2019/12/31", '%Y/%m/%d')
    cost_income_sheet['F10'].number_format = 'YYYY/MM/DD'
    cost_income_sheet['G10'] = "五年均值"

    cost_income_sheet['A11'] = "成本/收入"
    cost_income_sheet['A12'] = "销售费用/收入"
    cost_income_sheet['A13'] = "管理费用/收入"
    cost_income_sheet['A14'] = "研发费用/收入"
    cost_income_sheet['A15'] = "毛利润率=营业利润/收入"
    cost_income_sheet['A16'] = "所得税/利润总额"
































    row_1 = 2
    for i in range(5):
        calc_sheet[f'{years[i]}{row_1}'] = results_1['流动比率'][i]

    row_2 = 3
    for i in range(5):
        calc_sheet[f'{years[i]}{row_2}'] = results_2['速动比率'][i]

    row_3 = 4
    for i in range(5):
        calc_sheet[f'{years[i]}{row_3}'] = results_3['资产负债率'][i]
        # print(calc_sheet[f'{years[i]}{row_3}'])

    row_4 = 5
    for i in range(5):
        calc_sheet[f'{years[i]}{row_4}'] = results_4['长期资产适合率'][i]

    row_5 = 7
    for i in range(5):
        calc_sheet[f'{years[i]}{row_5}'] = results_5['资本保值增值率'][i]

    row_6 = 8
    for i in range(5):
        calc_sheet[f'{years[i]}{row_6}'] = results_6['资本积累率'][i]

    row_7 = 10
    for i in range(5):
        calc_sheet[f'{years[i]}{row_7}'] = results_income_1['主营业务毛利率'][i]
    row_8 = 11
    for i in range(5):
        calc_sheet[f'{years[i]}{row_8}'] = results_income_2['主营业务净利率'][i]
    row_9 = 12
    for i in range(5):
        calc_sheet[f'{years[i]}{row_9}'] = results_income_3['主营业务成本率'][i]
    row_10 = 13
    for i in range(5):
        calc_sheet[f'{years[i]}{row_10}'] = results_income_4['主营业务税金率'][i]
    row_11 = 14
    for i in range(5):
        calc_sheet[f'{years[i]}{row_11}'] = results_income_5['资产净利率'][i]
    row_12 = 15
    for i in range(5):
        calc_sheet[f'{years[i]}{row_12}'] = results_income_6['净资产收益率'][i]
    row_13 = 16
    for i in range(5):
        calc_sheet[f'{years[i]}{row_13}'] = results_income_7['销售费用率'][i]
    row_14 = 17
    for i in range(5):
        calc_sheet[f'{years[i]}{row_14}'] = results_income_8['管理费用率'][i]
    row_15 = 18
    for i in range(5):
        calc_sheet[f'{years[i]}{row_15}'] = results_income_9['研发费用率'][i]
    row_16 = 19
    for i in range(5):
        calc_sheet[f'{years[i]}{row_16}'] = results_income_10['财务费用率'][i]
    row_17 = 20
    for i in range(5):
        calc_sheet[f'{years[i]}{row_17}'] = results_income_11['成本、费用利润率'][i]
    row_18 = 21
    for i in range(5):
        calc_sheet[f'{years[i]}{row_18}'] = results_cash_1['销售收现比'][i]
    row_19 = 22
    for i in range(5):
        calc_sheet[f'{years[i]}{row_19}'] = results_cash_2['营运指数'][i]
    row_20 = 23
    for i in range(5):
        calc_sheet[f'{years[i]}{row_20}'] = results_cash_3['现金比率'][i]
    row_21 = 24
    for i in range(5):
        calc_sheet[f'{years[i]}{row_21}'] = results_cash_4['现金流动负债比'][i]
    row_22 = 25
    for i in range(5):
        calc_sheet[f'{years[i]}{row_22}'] = results_cash_5['现金债务总额比'][i]
    row_23 = 26
    for i in range(5):
        calc_sheet[f'{years[i]}{row_23}'] = results_cash_6['销售现金比率'][i]
    row_24 = 31
    for i in range(5):
        calc_sheet[f'{years[i]}{row_24}'] = results_cash_7['应收帐款周转率(次数)'][i]
    row_25 = 32
    for i in range(5):
        calc_sheet[f'{years[i]}{row_25}'] = results_cash_8['应收帐款周转天数'][i]
    row_26 = 33
    for i in range(5):
        calc_sheet[f'{years[i]}{row_26}'] = results_cash_9['存货周转率'][i]
    row_27 = 34
    for i in range(5):
        calc_sheet[f'{years[i]}{row_27}'] = results_cash_10['存货周转天数'][i]
    row_28 = 35
    for i in range(5):
        calc_sheet[f'{years[i]}{row_28}'] = results_cash_11['流动资产周转次数(率)'][i]
    row_29 = 36
    for i in range(5):
        calc_sheet[f'{years[i]}{row_29}'] = results_cash_12['流动资产周转天数'][i]
    row_30 = 37
    for i in range(5):
        calc_sheet[f'{years[i]}{row_30}'] = results_cash_13['总资产周转率'][i]
    row_31 = 38
    for i in range(5):
        calc_sheet[f'{years[i]}{row_31}'] = results_cash_14['总资产周转天数'][i]
    row_32 = 44
    for i in range(5):
        calc_sheet[f'{years[i]}{row_32}'] = results_cash_15['权益乘数'][i]
    row_33 = 45
    for i in range(5):
        calc_sheet[f'{years[i]}{row_33}'] = results_cash_16['销售净利率'][i]
    row_34 = 46
    for i in range(5):
        calc_sheet[f'{years[i]}{row_34}'] = results_cash_13['总资产周转率'][i]
    row_35 = 47
    for i in range(5):
        calc_sheet[f'{years[i]}{row_35}'] = results_cash_17['权益净利率'][i]
    row_36 = 63
    for i in range(5):
        calc_sheet[f'{years[i]}{row_36}'] = results_cash_4['现金流动负债比'][i]
    row_37 = 64
    for i in range(5):
        calc_sheet[f'{years[i]}{row_37}'] = results_cash_5['现金债务总额比'][i]
    row_38 = 67
    for i in range(5):
        calc_sheet[f'{years[i]}{row_38}'] = results_cash_6['销售现金比率'][i]
    row_39 = 69
    for i in range(5):
        calc_sheet[f'{years[i]}{row_39}'] = results_cash_18['全部资产现金回收率'][i]
    row_40 = 76
    for i in range(5):
        calc_sheet[f'{years[i]}{row_40}'] = results_income_4['主营业务税金率'][i]
    row_41 = 77
    for i in range(5):
        calc_sheet[f'{years[i]}{row_41}'] = results_income_12['所得税率'][i]

    row_42 = 2
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_42}'] = results_temp_operating_income_copy['营业总收入'][i]
    row_43 = 3
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_43}'] = results_cost_income_1['营业总成本'][i]

    row_44 = 4
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_44}'] = results_cost_income_2['营业税金及附加'][i]

    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_4}'] = results_cost_income_3['销售费用'][i]

    row_45 = 6
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_45}'] = results_cost_income_4['管理费用'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_5}'] = results_cost_income_5['研发费用'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_6}'] = results_cost_income_6['毛利润'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_8}'] = results_cost_income_7['成本/收入'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_9}'] = results_cost_income_8['销售费用/收入'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_10}'] = results_cost_income_9['管理费用/收入'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_11}'] = results_cost_income_10['研发费用/收入'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_12}'] = results_cost_income_11['毛利润率'][i]
    for i in range(5):
        cost_income_sheet[f'{years[i]}{row_13}'] = results_income_12['所得税率'][i]




































    for row in ['2','3','4','5','7','8','10','11','12','13','14','15','16','17','18','19','20',
                '21','22','23','24','25','26','31','33','35','37','44','45','46','47','63','64','67','69','76','77']:# many other rows
        for col in ['B', 'C', 'D', 'E', 'F']:
            cell = calc_sheet[f'{col}{row}']
            if cell.value != '-':
                cell.style = percent_style

    for row in ['11','12','13','14','15','16']:
        for col in ['B','C','D','E','F','G']:
            cell2 = cost_income_sheet[f'{col}{row}']
            if cell2.value != '-':
                cell2.style = percent_style


    right_alignment = Alignment(horizontal='right')
    cells_to_align_right = ['A1','B1','C1','D1','E1','F1',
                            'B2','C2','D2','E2','F2',
                            'B3','C3','D3','E3','F3',
                            'B4','C4','D4','E4','F4',
                            'B5','C5','D5','E5','F5',
                            'A6','B6','C6','D6','E6','F6',
                            'B7','C7','D7','E7','F7',
                            'B8','C8','D8','E8','F8',
                            'A9','B9','C9','D9','E9','F9',
                            'B10','C10','D10','E10','F10',
                            'B11','C11','D11','E11','F11',
                            'B12','C12','D12','E12','F12',
                            'B13','C13','D13','E13','F13',
                            'B14','C14','D14','E14','F14',
                            'B15','C15','D15','E15','F15',
                            'B16','C16','D16','E16','F16',
                            'B17','C17','D17','E17','F17',
                            'B18','C18','D18','E18','F18',
                            'B19','C19','D19','E19','F19',
                            'B20','C20','D20','E20','F20',
                            'B21','C21','D21','E21','F21',
                            'B22','C22','D22','E22','F22',
                            'B23','C23','D23','E23','F23',
                            'B24','C24','D24','E24','F24',
                            'B25','C25','D25','E25','F25',
                            'B26','C26','D26','E26','F26',
                            'B30','C30','D30','E30','F30',
                            'B31','C31','D31','E31','F31',
                            'B32','C32','D32','E32','F32',
                            'B33','C33','D33','E33','F33',
                            'B34','C34','D34','E34','F34',
                            'B35','C35','D35','E35','F35',
                            'B36','C36','D36','E36','F36',
                            'B37','C37','D37','E37','F37',
                            'B38','C38','D38','E38','F38',
                            'B43','C43','D43','E43','F43',
                            'B44','C44','D44','E44','F44',
                            'B45','C45','D45','E45','F45',
                            'B46','C46','D46','E46','F46',
                            'B47','C47','D47','E47','F47',
                            'B50','C50','D50','E50','F50',
                            'B62','C62','D62','E62','F62',
                            'B63','C63','D63','E63','F63',
                            'B64','C64','D64','E64','F64',
                            'B67','C67','D67','E67','F67',
                            'B69','C69','D69','E69','F69',
                            'B71','C71','D71','E71','F71',
                            'B76','C76','D76','E76','F76',
                            'B77','C77','D77','E77','F77']

    cells_to_align_right_cost_income = ['B1','C1','D1','E1','F1',
                                        'B2','C2','D2','E2','F2',
                                        'B3','C3','D3','E3','F3',
                                        'B4','C4','D4','E4','F4',
                                        'B5','C5','D5','E5','F5',
                                        'B6','C6','D6','E6','F6',
                                        'B7','C7','D7','E7','F7',
                                        'B8','C8','D8','E8','F8',
                                        'B10','C10','D10','E10',
                                        'F10','G1','G2','G3','G4','G10',
                                        'G5','G6','G7','G8','G11',
                                        'B11','C11','D11','E11','F11',
                                        'B12','C12','D12','E12','F12',
                                        'B13','C13','D13','E13','F13',
                                        'B14','C14','D14','E14','F14',
                                        'B15','C15','D15','E15','F15',
                                        'B16','C16','D16','E16','F16']

    for cell in cells_to_align_right:
        calc_sheet[cell].alignment = right_alignment

    for cell2 in cells_to_align_right_cost_income:
        cost_income_sheet[cell2].alignment = right_alignment



    if all(item != '-' for item in results_cost_income_7['成本/收入']):
        chart = AreaChart()

        for i in range(11, 16):  # 只添加第一和第二组数据
            values = Reference(cost_income_sheet, min_col=2, min_row=i, max_col=6, max_row=i)
            series = Series(values, title=cost_income_sheet[f'A{i}'].value)
            chart.series.append(series)

        categories = Reference(cost_income_sheet, min_col=2, max_col=6, min_row=10, max_row=10)
        chart.set_categories(categories)

        chart.title = "成本收入比率"

        chart.legend.position = 'b'

        cost_income_sheet.add_chart(chart, "H1")

    workbook.save(file_name)





