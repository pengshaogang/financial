import copy
import os
import openpyxl

from openpyxl.styles import Alignment
from openpyxl.styles import NamedStyle






file_names = [f for f in os.listdir('.') if f.startswith('汽车')]

columns_to_delete = ['K', 'I', 'G', 'E', 'C']

for file_name in file_names:
    workbook = openpyxl.load_workbook(file_name)
    cal_sheet_name = [s for s in workbook.sheetnames if "计算公式" in s]

    for sheet_name in cal_sheet_name:
        sheet_cal = workbook[sheet_name]
        first_row = list(sheet_cal[1])  # 行索引在 openpyxl 中是从1开始的，所以第二行是2
        # 遍历第二行的单元格
        for cell in first_row:
            if 'ave' in str(cell.value).lower():  # 检查单元格中是否包含 'ave'
                # 删除包含 'ave' 的列
                sheet_cal.delete_cols(cell.column)  # cell.column 为该单元格所在列的列号

    # 保存修改
    workbook.save(file_name)




values_b_1, values_c_1, values_d_1, values_e_1, values_f_1 = [], [], [], [], []#1代表第一个指标，b,c,d,e,f代表原先的单元格

rows_to_process = [2,3,4,5, 7,8]

rows_to_process.extend(range(10, 27))

rows_to_process.extend(range(31,39))

rows_to_process.extend(range(44,48))

rows_to_process.extend(range(63,65))
rows_to_process.extend(range(67,68))
rows_to_process.extend(range(69,70))
rows_to_process.extend(range(76,78))

columns = ['B', 'C', 'D', 'E', 'F']
keys = ['val1', 'val2', 'val3','val4', 'val6','val7']

keys.extend([f'val{i}' for i in range(9, 26)])  # 使用列表推导来生成和添加键
keys.extend([f'val{i}' for i in range(30,38)])
keys.extend([f'val{i}' for i in range(43,47)])
keys.extend([f'val{i}' for i in range(62,64)])
keys.extend([f'val{i}' for i in range(66,67)])
keys.extend([f'val{i}' for i in range(68,69)])
keys.extend([f'val{i}' for i in range(75,77)])


# 使用字典推导和循环初始化values字典
values = {col: {key: [] for key in keys} for col in columns}


for file_name in file_names:
    workbook = openpyxl.load_workbook(file_name)
    cal_sheet_name = [s for s in workbook.sheetnames if "计算公式" in s]

    for sheet_name in cal_sheet_name:
        sheet_cal = workbook[sheet_name]
        # 读取B2单元格的值并添加到列表中
        # 读取B2到F2单元格的值并添加到相应的列表中

        for row_num in rows_to_process:
            for col, value_list in zip(['B', 'C', 'D', 'E', 'F'],
                                       [values_b_1, values_c_1, values_d_1, values_e_1, values_f_1]):
                cell_value = sheet_cal[f'{col}{row_num}'].value  # 动态使用row_num来访问行号
                values[col][f'val{row_num - 1}'].append(cell_value)








ave = {}  # 用于存储平均值的字典
for col in columns:
    ave[col] = {}
    for key in keys:
        ave[col][key] = []
        value_list1 = values[col][key]
        if '-' in value_list1:
            ave[col][key].append('-')  # 包含'-'则在平均值列表中添加'-'
        else:
            if value_list1:  # 确保列表非空
                # 计算平均值并添加到对应的列的平均值列表中
                average_value = float(sum(value_list1)) / float(len(value_list1))
                ave[col][key].append(average_value)
            else:
                print("something wrong!")  # 如果数据列表为空

print(len(value_list1))
print(ave)


percent_style_name = 'percent_style'

# 检查样式是否已存在
existing_styles = [style for style in workbook.named_styles]
if percent_style_name not in existing_styles:
    percent_style = NamedStyle(name=percent_style_name, number_format='0.00%')
    workbook.add_named_style(percent_style)
else:
    percent_style = next(style for style in workbook.named_styles if style == percent_style_name)


right_alignment = Alignment(horizontal='right')


#保存到.xlsx

column_map = {
    'B': 'C',
    'C': 'E',
    'D': 'G',
    'E': 'I',
    'F': 'K'
}

for file_name in file_names:
    workbook = openpyxl.load_workbook(file_name)
    cal_sheet_names = [s for s in workbook.sheetnames if "计算公式" in s]
    for sheet_name in cal_sheet_names:
        cal_sheet = workbook[sheet_name]
        # 在B, C, D, E, F列的右边分别添加新列，并设置第二行的值为对应的平均值
        original_columns = ['F', 'E', 'D', 'C', 'B']
        for col in original_columns:
            col_index = openpyxl.utils.column_index_from_string(col) + 1
            cal_sheet.insert_cols(col_index)  # 插入新列

        for col, key_vals in ave.items():
            for key, value in key_vals.items():
                # 从key（如'val1'）中解析出行号，行号为key中数字+1
                if col in column_map:
                    new_col = column_map[col]
                    new_col_index = openpyxl.utils.column_index_from_string(new_col)
                    row_num = int(key[3:]) + 1  # key的形式是'val1'，其中'1'是数字部分


                    # 设置列的标题为'ave'在第一行
                    cal_sheet.cell(row=1,  column=new_col_index).value = 'ave'
                    cal_sheet.cell(row=30, column=new_col_index).value = 'ave'
                    cal_sheet.cell(row=43, column=new_col_index).value = 'ave'
                    cal_sheet.cell(row=62, column=new_col_index).value = 'ave'
                    cal_sheet.cell(row=66, column=new_col_index).value = 'ave'
                    cal_sheet.cell(row=75, column=new_col_index).value = 'ave'
                    # 将平均值设置在对应的行和列
                    cal_sheet.cell(row=row_num, column=new_col_index).value = value[0]


    cal_sheet.column_dimensions['H'].width = 20
    cal_sheet.column_dimensions['I'].width = 20
    cal_sheet.column_dimensions['J'].width = 20
    cal_sheet.column_dimensions['K'].width = 20

    for row in rows_to_process:
        if row == 32 or row == 34 or row == 36 or row == 38:
            continue
        for col in ['C','E','G','I','K']:
            cell1 = cal_sheet[f'{col}{row}']
            if cell1.value != '-':
                cell1.style = percent_style

    cells_to_align_right = ['C1','E1','G1','I1','K1',
                            'C5','E5','G5','I5','K5',
                            'C7','E7','G7','I7','K7',
                            'C8','E8','G8','I8','K8',
                            'C14','E14','G14','I14','K14',
                            'C15','E15','G15','I15','K15',
                            'C30','E30','G30','I30','K30',
                            'C31', 'E31', 'G31', 'I31', 'K31',
                            'C32', 'E32', 'G32', 'I32', 'K32',
                            'C33', 'E33', 'G33', 'I33', 'K33',
                            'C34', 'E34', 'G34', 'I34', 'K34',
                            'C35', 'E35', 'G35', 'I35', 'K35',
                            'C36', 'E36', 'G36', 'I36', 'K36',
                            'C37', 'E37', 'G37', 'I37', 'K37',
                            'C38', 'E38', 'G38', 'I38', 'K38',
                            'C43', 'E43', 'G43', 'I43', 'K43',
                            'C46', 'E46', 'G46', 'I46', 'K46',
                            'C47', 'E47', 'G47', 'I47', 'K47',
                            'C62', 'E62', 'G62', 'I62', 'K62',
                            'B66', 'D66', 'F66', 'H66', 'J66',
                            'C66', 'E66', 'G66', 'I66', 'K66',
                            'C75', 'E75', 'G75', 'I75', 'K75',
                            'C77', 'E77', 'G77', 'I77', 'K77'
                            ]

    for cell in cells_to_align_right:
        cal_sheet[cell].alignment = right_alignment


    workbook.save(file_name)  # 保存修改后的工作簿





